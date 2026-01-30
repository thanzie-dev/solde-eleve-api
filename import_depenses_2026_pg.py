import psycopg
from openpyxl import load_workbook
from datetime import datetime
import os
import csv

# =====================================================
# CONFIGURATION
# =====================================================

EXCEL_FILE = "DEPENSES_2026.xlsx"
ERROR_FILE = "erreurs_import_depenses.csv"
SHEET_NAME = "Feuil1"
START_ROW = 13

MAX_EMPTY_DATES = 20

# Colonnes (Excel commence réellement en colonne B)
COL_REF_DP   = 3   # C
COL_DATE     = 4   # D
COL_REPORT   = 5   # E
COL_BLOC1    = 6   # F
COL_BLOC2    = 7   # G
COL_BUS1     = 8   # H
COL_BUS2     = 9   # I
COL_TOT_ENTR = 10  # J (calculée Excel, NON importée)
COL_LB_DP    = 11  # K
COL_MT_DP    = 12  # L
COL_BANQUE   = 13  # M
COL_SOLDE    = 14  # N (calculée Excel, NON importée)
COL_LB_OBS   = 15  # O
COL_TT_OBS   = 16  # P
COL_ANNEE    = 17  # Q

DATABASE_URL = os.environ.get("DATABASE_URL")

# =====================================================
# OUTILS DE NETTOYAGE / VALIDATION
# =====================================================

def to_float_checked(value, row, col_name, errors):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = (
            value.replace("\u00a0", "")
                 .replace(" ", "")
                 .replace(",", ".")
                 .strip()
        )
        try:
            return float(cleaned)
        except ValueError:
            errors.append((row, col_name, value, "Nombre invalide"))
            return None
    errors.append((row, col_name, value, "Type numérique invalide"))
    return None


def parse_date_checked(value, row, errors):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d/%m/%Y").date()
        except ValueError:
            errors.append((row, "DATE", value, "Format date invalide (JJ/MM/AAAA)"))
            return None
    errors.append((row, "DATE", value, "Type de date invalide"))
    return None


# =====================================================
# LECTURE EXCEL
# =====================================================

caisse_rows = []
depense_rows = []
obs_rows = []
errors = []

wb = load_workbook(EXCEL_FILE, data_only=True)
ws = wb[SHEET_NAME]

empty_date_count = 0
caisse_dates_importees = set()

for row in range(START_ROW, ws.max_row + 1):

    date_cell = ws.cell(row=row, column=COL_DATE).value
    date_op = parse_date_checked(date_cell, row, errors)

    if not date_op:
        empty_date_count += 1
        if empty_date_count >= MAX_EMPTY_DATES:
            break
        continue
    else:
        empty_date_count = 0

    annee = ws.cell(row=row, column=COL_ANNEE).value
    if not annee:
        errors.append((row, "ANNEE SCOLAIRE", None, "Année scolaire manquante"))
        continue

    # =================================================
    # CAISSE JOURNALIÈRE (UNE SEULE FOIS PAR JOUR)
    # =================================================

    key = (date_op, annee)

    if key not in caisse_dates_importees:
        vals = {}

        for name, col in {
            "REPORT": COL_REPORT,
            "BLOC1": COL_BLOC1,
            "BLOC2": COL_BLOC2,
            "BUS1": COL_BUS1,
            "BUS2": COL_BUS2
        }.items():
            v = to_float_checked(ws.cell(row=row, column=col).value, row, name, errors)
            if v is None:
                break
            vals[name] = v
        else:
            caisse_rows.append((
                date_op,
                vals["REPORT"],
                vals["BLOC1"],
                vals["BLOC2"],
                vals["BUS1"],
                vals["BUS2"],
                annee
            ))
            caisse_dates_importees.add(key)

    # =================================================
    # DÉPENSE (FIDÈLE À LA FORMULE EXCEL)
    # Dépense réelle = MT DEP + BANQUE
    # =================================================

    mt_dep = to_float_checked(ws.cell(row=row, column=COL_MT_DP).value, row, "MT DEP", errors)
    banque = to_float_checked(ws.cell(row=row, column=COL_BANQUE).value, row, "BANQUE", errors)

    if mt_dep is not None and banque is not None:
        if (mt_dep + banque) > 0:
            depense_rows.append((
                ws.cell(row=row, column=COL_REF_DP).value,
                date_op,
                ws.cell(row=row, column=COL_LB_DP).value,
                mt_dep,
                banque,
                annee
            ))

    # =================================================
    # OBSERVATIONS
    # =================================================

    tt_obs = to_float_checked(ws.cell(row=row, column=COL_TT_OBS).value, row, "TT OBS", errors)
    if tt_obs is not None and tt_obs > 0:
        obs_rows.append((
            date_op,
            ws.cell(row=row, column=COL_LB_OBS).value,
            tt_obs,
            annee
        ))


# =====================================================
# RAPPORT D’ERREURS
# =====================================================

if errors:
    with open(ERROR_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Ligne Excel", "Colonne", "Valeur", "Erreur"])
        writer.writerows(errors)


# =====================================================
# INSERT EN BASE (ROBUSTE)
# =====================================================

with psycopg.connect(DATABASE_URL) as conn:
    with conn.cursor() as cur:

        if caisse_rows:
            cur.executemany("""
                INSERT INTO caisse_journaliere
                (date_operation, report, bloc1, bloc2, bus1, bus2, annee_scolaire)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (date_operation, annee_scolaire)
                DO UPDATE SET
                    report = EXCLUDED.report,
                    bloc1  = EXCLUDED.bloc1,
                    bloc2  = EXCLUDED.bloc2,
                    bus1   = EXCLUDED.bus1,
                    bus2   = EXCLUDED.bus2
            """, caisse_rows)

        if depense_rows:
           cur.executemany("""
                INSERT INTO depense
                (ref_dp, date_depense, libelle, montant, banque, annee_scolaire)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON CONFLICT DO NOTHING
            """, depense_rows)


        if obs_rows:
            cur.executemany("""
                INSERT INTO observation
                (date_operation, libelle, montant, annee_scolaire)
                VALUES (%s,%s,%s,%s)
                ON CONFLICT (date_operation, libelle, annee_scolaire)
                DO NOTHING
            """, obs_rows)

        conn.commit()


# =====================================================
# RÉSUMÉ
# =====================================================

print(" Import terminé (version corrigée & fidèle Excel)")
print(f"   - Caisse journalière : {len(caisse_rows)} lignes")
print(f"   - Dépenses : {len(depense_rows)} lignes")
print(f"   - Observations : {len(obs_rows)} lignes")

if errors:
    print(f" {len(errors)} erreurs détectées")
    print(f" Voir : {ERROR_FILE}")
else:
    print(" Aucune erreur détectée")
