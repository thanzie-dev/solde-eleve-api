"""
import_excel_pg.py — VERSION STRICTE MÉTIER (DATE OBLIGATOIRE)
✔ psycopg v3
✔ Sans pandas
✔ openpyxl pur
✔ Chaque paiement garde SA date
✔ Import BLOQUÉ si DatePaiement absente/invalide
✔ Bug Excel date corrigé (respect du type date)
✔ Compatible Python 3.13 / Render / Local
"""

import os
import sys
import re
from datetime import datetime, date

import psycopg
from openpyxl import load_workbook

# ======================================================
# CONFIGURATION
# ======================================================

EXCEL_FILE = "THZBD2526GA.xlsx"
DATABASE_URL = os.environ.get("DATABASE_URL")
HEADER_LINE = 8

if not DATABASE_URL:
    raise RuntimeError("❌ DATABASE_URL non définie")

REQUIRED_COLS = [
    "Matricule", "Nom", "Sexe", "Classe", "Categorie", "Section",
    "Telephone", "Email", "NumRecu", "Mois",
    "FIP", "FF", "Obs", "Jour", "DatePaiement", "AnneeScolaire"
]

# ======================================================
# OUTILS
# ======================================================

def log(msg):
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}")

def clean(v):
    """Nettoyage standard (SAUF pour les dates)"""
    if v is None:
        return None
    s = str(v).strip()
    return s or None

def to_float(v):
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return 0.0

def parse_date(v):
    """
    Analyse robuste des dates :
    - date Excel (date)
    - date Excel (datetime)
    - chaîne '06/11/2025', '2025-11-06', '06-11-2025'
    """
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                pass
    return None

def matricule_valide(m):
    return bool(m and re.match(r"^(PL|LT)\d+$", str(m)))

# ======================================================
# LECTURE & VALIDATION STRICTE DE L'EXCEL
# ======================================================

def charger_excel_strict():
    log("Lecture et validation stricte du fichier Excel…")

    wb = load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    ws = wb.active

    headers = [clean(c.value) for c in ws[HEADER_LINE]]
    col = {h: i for i, h in enumerate(headers) if h}

    missing = set(REQUIRED_COLS) - set(col)
    if missing:
        raise RuntimeError(f"❌ Colonnes manquantes : {missing}")

    lignes = []

    for idx, row in enumerate(
        ws.iter_rows(min_row=HEADER_LINE + 1, values_only=True),
        start=HEADER_LINE + 1
    ):
        r = {}

        # ⚠️ IMPORTANT : NE PAS nettoyer DatePaiement ici
        for k in REQUIRED_COLS:
            cell_value = row[col[k]]
            if k == "DatePaiement":
                r[k] = cell_value   # on garde le type Excel
            else:
                r[k] = clean(cell_value)

        if not matricule_valide(r["Matricule"]):
            continue
        if not r["NumRecu"]:
            continue

        # 🔴 VALIDATION DATE STRICTE
        date_paiement = parse_date(r["DatePaiement"])
        if date_paiement is None:
            raise RuntimeError(
                f"\n❌ IMPORT BLOQUÉ\n"
                f"DatePaiement manquante ou invalide\n"
                f"Ligne Excel : {idx}\n"
                f"Matricule   : {r['Matricule']}\n"
                f"NumRecu     : {r['NumRecu']}\n"
                f"Valeur brute DatePaiement : {r['DatePaiement']}\n"
                f"👉 Corrigez le fichier Excel puis relancez l’import.\n"
            )

        r["DatePaiement"] = date_paiement
        r["FIP"] = to_float(r["FIP"])
        r["FF"] = to_float(r["FF"])

        lignes.append(r)

    log(f"{len(lignes)} lignes valides prêtes pour import")
    return lignes

# ======================================================
# INSERTION TRANSACTIONNELLE (PSYCOPG3)
# ======================================================

def inserer_donnees(lignes, conn):
    conn.autocommit = False

    with conn.cursor() as cur:

        # ---------- ÉLÈVES ----------
        eleves = {
            r["Matricule"]: (
                r["Matricule"], r["Nom"], r["Sexe"], r["Classe"],
                r["Categorie"], r["Section"], r["Telephone"], r["Email"]
            )
            for r in lignes
        }

        cur.executemany("""
            INSERT INTO eleves (
                matricule, nom, sexe, classe,
                categorie, section, telephone, email
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (matricule) DO UPDATE SET
                nom=EXCLUDED.nom,
                sexe=EXCLUDED.sexe,
                classe=EXCLUDED.classe,
                categorie=EXCLUDED.categorie,
                section=EXCLUDED.section,
                telephone=EXCLUDED.telephone,
                email=EXCLUDED.email;
        """, eleves.values())

        # ---------- MAP matricule → eleve_id ----------
        cur.execute("SELECT id, matricule FROM eleves")
        eleve_ids = {m: i for i, m in cur.fetchall()}

        # ---------- PAIEMENTS ----------
        paiements = []
        for r in lignes:
            eid = eleve_ids.get(r["Matricule"])
            if not eid:
                continue

            paiements.append((
                eid,
                r["NumRecu"],
                r["Mois"],
                r["FIP"],
                r["FF"],
                r["Obs"],
                r["Jour"],
                r["DatePaiement"],
                r["AnneeScolaire"]
            ))

        cur.executemany("""
            INSERT INTO paiements (
                eleve_id, numrecu, mois, fip, ff,
                obs, jour, datepaiement, annee_scolaire
            )
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (numrecu) DO UPDATE SET
                mois=EXCLUDED.mois,
                fip=EXCLUDED.fip,
                ff=EXCLUDED.ff,
                obs=EXCLUDED.obs,
                jour=EXCLUDED.jour,
                datepaiement=EXCLUDED.datepaiement,
                annee_scolaire=EXCLUDED.annee_scolaire;
        """, paiements)

    conn.commit()

# ======================================================
# MAIN
# ======================================================

def run_import():
    log("=== DÉBUT IMPORT (MODE STRICT DATE) ===")

    lignes = charger_excel_strict()

    with psycopg.connect(
        DATABASE_URL,
        sslmode="require" if "render.com" in DATABASE_URL else "disable"
    ) as conn:
        inserer_donnees(lignes, conn)

    log("✅ IMPORT TERMINÉ AVEC SUCCÈS")

if __name__ == "__main__":
    try:
        run_import()
    except Exception as e:
        log("❌ IMPORT ANNULÉ")
        print(e)
        sys.exit(1)
