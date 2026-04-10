"""
IMPORT INSCRIPTION — VERSION PRODUCTION V5 (STABLE)

✔ 1 seule table (inscription)
✔ 19 colonnes respectées
✔ Anti-doublons (numrecu)
✔ Nettoyage Excel robuste
✔ Gestion catégories + section
✔ Compatible Flask (importer_inscriptions)
"""

import os
import sys
import re
import io
import csv
from datetime import datetime, date

import psycopg
from openpyxl import load_workbook

# ======================================================
# CONFIG
# ======================================================

EXCEL_FILE = "INSC_THZ2526.xlsx"
DATABASE_URL = os.environ.get("DATABASE_URL")
HEADER_LINE = 8

if not DATABASE_URL:
    raise RuntimeError("❌ DATABASE_URL non définie")

REQUIRED_COLS = [
    "Num", "Matricule", "NumRecu", "Telephone", "Sexe", "Categorie",
    "Nom", "Classe", "Finsc", "Jour", "Mois", "DateInsc",
    "Adresse", "Obs", "LieuDnss", "Respo", "AnneeScolaire",
    "Section", "Email"
]

CATEGORIES_VALIDES = {"PY", "NPY", "ABD"}

# ======================================================
# OUTILS
# ======================================================

def log(msg):
    print(f"[{datetime.now():%Y-%m-%d %H:%M:%S}] {msg}")
    
def log_import(nb, statut, msg=""):
    with psycopg.connect(DATABASE_URL) as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO import_log (nb_lignes, statut, commentaire)
                VALUES (%s, %s, %s)
            """, (nb, statut, msg))
        conn.commit()  
    


def clean(v):
    if v is None:
        return None
    return str(v).replace("\n", "").replace("\r", "").strip() or None

def to_float(v):
    try:
        return float(str(v).replace(",", "."))
    except:
        return 0.0

def safe_int(v):
    try:
        return int(float(str(v)))
    except:
        return None

def parse_date(v):
    if v is None:
        return None

    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, date):
        return v

    if isinstance(v, str):
        v = v.strip().split(" ")[0]

        formats = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]

        for f in formats:
            try:
                return datetime.strptime(v, f).date()
            except:
                pass

    return None

def matricule_valide(m):
    return bool(m and re.match(r"^(PL|LT)\d+$", str(m)))

def normaliser_categorie(val):
    if not val:
        return "NPY"

    v = str(val).strip().upper()

    mapping = {
        "PAYANT": "PY",
        "P": "PY",
        "NP": "NPY",
        "NON PAYANT": "NPY",
        "AB": "ABD",
        "ABANDON": "ABD"
    }

    v = mapping.get(v, v)
    return v if v in CATEGORIES_VALIDES else "NPY"


def normaliser_section(val):
    if not val:
        return "PRM"

    v = str(val).strip().upper()

    # Nettoyage avancé
    v = v.replace(".", "").replace("-", "").strip()

    mapping = {
        # PRIMAIRE
        "PRIMAIRE": "PRM",
        "PRM": "PRM",
        "P": "PRM",
        "PM": "PRM",

        # SECONDAIRE
        "SECONDAIRE": "SEC",
        "SEC": "SEC",
        "S": "SEC",

        # MATERNELLE
        "MATERNELLE": "MAT",
        "MAT": "MAT",
        "M": "MAT",
        "MT": "MAT"
    }

    # 🔥 fallback intelligent
    if v not in mapping:
        if "MAT" in v:
            return "MAT"
        if "SEC" in v:
            return "SEC"
        if "PRI" in v or "PRM" in v:
            return "PRM"

    return mapping.get(v, "PRM")



# ======================================================
# DB INIT
# ======================================================

def init_db(conn):
    with conn.cursor() as cur:

        cur.execute("""
        CREATE TABLE IF NOT EXISTS inscription (
            id SERIAL PRIMARY KEY,
            num INTEGER,
            matricule VARCHAR(50),
            numrecu VARCHAR(50) UNIQUE,
            telephone VARCHAR(20),
            sexe VARCHAR(10),
            categorie VARCHAR(10),
            nom VARCHAR(100),
            classe VARCHAR(50),
            finsc NUMERIC,
            jour INTEGER,
            mois INTEGER,
            dateinsc DATE,
            adresse TEXT,
            obs TEXT,
            lieudnss TEXT,
            respo TEXT,
            annee_scolaire TEXT,
            section VARCHAR(10),
            email TEXT
        );
        """)

    conn.commit()

# ======================================================
# LECTURE EXCEL
# ======================================================

def charger_excel():
    log("Lecture Excel...")

    wb = load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    ws = wb.active

    headers = [clean(c.value) for c in ws[HEADER_LINE]][1:]
    col = {h: i for i, h in enumerate(headers) if h}

    missing = set(REQUIRED_COLS) - set(col)
    if missing:
        raise RuntimeError(f"❌ Colonnes manquantes : {missing}")

    lignes = []

    matricules = set()
    numrecus = set()

    erreurs = []

    for idx, row in enumerate(ws.iter_rows(min_row=HEADER_LINE + 1, values_only=True), start=HEADER_LINE + 1):

        row = row[1:]
        r = {k: clean(row[col[k]]) for k in REQUIRED_COLS}

        # ========= VALIDATION =========

        # Matricule obligatoire
        if not r["Matricule"]:
            erreurs.append(f"Ligne {idx} → Matricule vide")
            continue

        # NumRecu obligatoire
        if not r["NumRecu"] or r["NumRecu"] == "0":
            erreurs.append(f"Ligne {idx} → NumRecu invalide (0 ou vide)")
            continue

        # Doublon matricule
        if r["Matricule"] in matricules:
            erreurs.append(f"Ligne {idx} → Doublon matricule : {r['Matricule']}")
            continue

        # Doublon numrecu
        if r["NumRecu"] in numrecus:
            erreurs.append(f"Ligne {idx} → Doublon NumRecu : {r['NumRecu']}")
            continue

        matricules.add(r["Matricule"])
        numrecus.add(r["NumRecu"])

        # ========= TRANSFORMATION =========
        r["Jour"] = safe_int(r["Jour"])
        r["Mois"] = safe_int(r["Mois"])
        r["Finsc"] = to_float(r["Finsc"])

        r["Categorie"] = normaliser_categorie(r["Categorie"])
        r["Section"] = normaliser_section(r["Section"])

        if not matricule_valide(r["Matricule"]):
            erreurs.append(f"Ligne {idx} → Matricule invalide")
            continue

        d = parse_date(r["DateInsc"])
        if not d:
            erreurs.append(f"Ligne {idx} → Date invalide")
            continue

        r["DateInsc"] = d

        lignes.append(r)

    # ========= BLOQUER SI ERREURS =========
    if erreurs:
        log("❌ ERREURS DÉTECTÉES — IMPORT BLOQUÉ")
        for e in erreurs[:20]:
            print(" -", e)

        raise RuntimeError(f"{len(erreurs)} erreurs détectées dans le fichier Excel")

    log(f"{len(lignes)} lignes valides")
    return lignes

# ======================================================
# INSERTION
# ======================================================

def inserer_donnees_copy(lignes, conn):
   

    with conn.cursor() as cur:

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        for r in lignes:
            writer.writerow([
                safe_int(r["Num"]),
                r["Matricule"],
                r["NumRecu"],
                r["Telephone"],
                r["Sexe"],
                r["Categorie"],
                r["Nom"],
                r["Classe"],
                r["Finsc"],
                r["Jour"],
                r["Mois"],
                r["DateInsc"],
                r["Adresse"],
                r["Obs"],
                r["LieuDnss"],
                r["Respo"],
                r["AnneeScolaire"],
                r["Section"],
                r["Email"]
            ])

        buffer.seek(0)

        # ✅ VERSION PSYCOPG V3
        with cur.copy("""
            COPY inscription (
                num, matricule, numrecu, telephone, sexe,
                categorie, nom, classe,
                finsc, jour, mois, dateinsc,
                adresse, obs, lieudnss, respo,
                annee_scolaire, section, email
            )
            FROM STDIN WITH (FORMAT CSV)
        """) as copy:

            copy.write(buffer.read())

    conn.commit()

# ======================================================
# EXPORT FLASK
# ======================================================

def importer_inscriptions():

    log("=== IMPORT INSCRIPTION ===")

    try:
        # 1. Charger Excel
        lignes = charger_excel()

        # 2. Nettoyage
        log("Nettoyage table...")
        with psycopg.connect(DATABASE_URL) as conn:
            with conn.cursor() as cur:
                cur.execute("TRUNCATE TABLE inscription RESTART IDENTITY CASCADE;")
            conn.commit()

        # 3. Import rapide
        log("Insertion des données...")
        with psycopg.connect(DATABASE_URL) as conn:
            inserer_donnees_copy(lignes, conn)

        # 4. Historique succès
        log_import(len(lignes), "SUCCES")

        log("✅ IMPORT TERMINÉ")

    except Exception as e:
        # 5. Historique erreur
        log_import(0, "ECHEC", str(e))

        log("❌ IMPORT ÉCHOUÉ")
        print(e)
        raise

# ======================================================
# MAIN
# ======================================================

if __name__ == "__main__":
    try:
        importer_inscriptions()
    except Exception as e:
        log("❌ IMPORT ÉCHOUÉ")
        print(e)
        sys.exit(1)